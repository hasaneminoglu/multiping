def ping_excel(siteadress=None):
    #siteadress="dd.de"
    import os
    import platform
    import matplotlib.pyplot as plt
    import xlsxwriter
    import pandas as pd
    import numpy as np
    from openpyxl import load_workbook
    from openpyxl.chart import LineChart, Reference
    import time

    plat = platform.system()
    os.chdir('C:\GRK\logs\multiping\\')

    test_nerede = open(siteadress+".txt", "r", encoding="utf-8")
    test_dizin = test_nerede.readlines()
    test_nerede.close()
    #print(test_dizin)
    ping_sonuclar_list = []

    sayı = 1
    for i in test_dizin:

        protokol = i.split("_")[8]
        list_zaman_list = []
        #print(sayı)
        sayı += 1
        #print(i)
        path = r"" + str(i.rstrip("\n"))
        os.chdir(path)
        test_sonuclar_file = open("multi_test_islem.txt", "r", encoding="utf-8")
        test_sonuclar = test_sonuclar_file.readlines()
        # print(test_sonuclar)
        test_sonuclar_file.close()
        hangi_setup_pc = test_sonuclar[0].split("_")[6]
        siteadress = test_sonuclar[0].split("_")[7]
        kac_kez = test_sonuclar[0].split("_")[9]
        # print(hangi_setup_pc,siteadress,kac_kez)
        sonuc_excel = r'' + hangi_setup_pc + '_' + siteadress + '_' + kac_kez + '_' + 'zaman' + '.xlsx'
        if not os.path.exists(sonuc_excel):
            workbook = xlsxwriter.Workbook(sonuc_excel)
            workbook.close()
        workbook = load_workbook(sonuc_excel)
        #print("test_sonuclar",test_sonuclar)
        for j in test_sonuclar:
            #print("jjjjj",j)
            if len(j.split(" : ")) > 2:
                sonuclar = j.split(" : ")[-1].strip("\n")
                sonuclar_list = sonuclar.split("_")
                #print(sonuclar_list)
                kanal= sonuclar_list[5]
                hangi_setup_pc = sonuclar_list[6]
                siteadress = sonuclar_list[7]
                ipadress = sonuclar_list[8]
                kac_kez = sonuclar_list[9]
                #print("*************")
                #print(hangi_setup_pc,siteadress,ipadress,kac_kez)
                ReadFile = open(sonuclar, "r")
                for line in ReadFile:
                    line = line.rstrip()
                    # print(line)
                ReadFile.close()
                ReadFile = open(sonuclar, "r")
                zaman_list = []
                for line in ReadFile:
                    if line.find("time=") != -1:
                        time = line.split("time=", 1)[1]
                        zaman = time.split("ms", 1)[0]
                        if plat == "Windows":
                            zaman_list.append(int(zaman))
                        else:
                            zaman_list.append(float(zaman))
                    elif line.find("Request") != -1:
                        zaman_list.append(int(5000))
                    elif line.find("Destination") != -1:
                        zaman_list.append(int(10000))
                    elif line.find("PING: transmit failed") != -1:
                        zaman_list.append(int(10000))
                    elif line.find("General failure") != -1:
                        zaman_list.append(int(10000))
                ReadFile.close()

                request_matched_indexes = []
                general_matched_indexes = []

                list_zaman_list.append(zaman_list)

                i = 0
                length = len(zaman_list)

                while i < length:
                    if (5000) == zaman_list[i]:
                        request_matched_indexes.append(i + 1)
                    elif (10000) == zaman_list[i]:
                    elif zaman_list[i] < (1):
                        zaman_list[i] = 1
                    i += 1

                #print(request_matched_indexes, general_matched_indexes)
                #print(len(general_matched_indexes))

                request_oranı = float(len(request_matched_indexes) / len(zaman_list)) * 100
                request_oranı_str = str(round(request_oranı, 2))
                general_oranı = float(len(general_matched_indexes) / len(zaman_list)) * 100
                general_oranı_str = str(round(general_oranı, 2))
                #kac_kez_general_oranı = str(len(general_matched_indexes))

                saniye = range(1, len(zaman_list) + 1)
                maxY = max(zaman_list)
                #print(zaman_list)
                """
                a = np.array(zaman_list)
                print(a)
                """
                data = {'Time response süresi': (zaman_list),
                        'Hangi pingte Request Timeout Olmuş': (request_matched_indexes),
                        'Hangi pingte General Failure': (general_matched_indexes),
                        'Request Timeout Oranı': [request_oranı],
                        'General Failure Oranı': [general_oranı]}
                writer = pd.ExcelWriter(sonuc_excel, engine='openpyxl')

                writer.book = workbook

                df = pd.DataFrame.from_dict(data, orient="index")
                columns = ['Time response süresi', 'Hangi pingte Request Timeout Olmuş', 'Hangi pingte General Failure',
                           'Request Timeout Oranı', 'General Failure Oranı']
                #print(df)
                #print(df.transpose())
                df = df.transpose()

                df.index = np.arange(1, len(df) + 1)
                df.index.name = "Hangi Timeout Ping"
                # df.columns=["time response süresi"]
                df.to_excel(writer, sheet_name=ipadress)

                ws = workbook[ipadress]
                #print(ipadress)
                # ws = workbook.active
                values = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=(len(zaman_list) + 1))
                chart = LineChart()
                chart.add_data(values)
                chart.x_axis.title = 'hangi saniye'
                uzunluk=len(zaman_list)/5
                chart.x_axis.tickLblSkip = uzunluk
                chart.y_axis.title = 'time response (ms)'
                chart.legend = None
                chart.width = 20
                ws.add_chart(chart, "H2")

                sheet_list = workbook.sheetnames
                for i in sheet_list:
                    if i == "Sheet1":
                        workbook.remove(workbook["Sheet1"])
                writer.save()
                writer.close()
                ping_sonuc_list = []
                ping_sonuc_list.append(ipadress)
                ping_sonuc_list.append(protokol)
                ping_sonuc_list.append(request_oranı_str)
                ping_sonuc_list.append(general_oranı_str)
                ping_sonuc_list.append(len(general_matched_indexes))
                ping_sonuc_list.append(kanal)

                ping_sonuclar_list.append(ping_sonuc_list)

        plot_list_zaman = []

        #print(plot_list_zaman)
        for i in list_zaman_list:
            plot_list = []
            # print(i)
            plot_list.append(saniye)
            plot_list.append(i)
            plot_list_zaman.append(plot_list)
        #print(plot_list_zaman)
        #print("--------------------------------")
        fig, ax = plt.subplots()
        for x, y in plot_list_zaman:
            ax.plot(x, y)
        plt.xlabel('hangi saniye')
        plt.ylabel('time response (ms)')
        plt.ylim(ymin=0)
        F = plt.gcf()
        Size = F.get_size_inches()
        F.set_size_inches(Size[1] * 4, Size[1] * 1, forward=True)
        plt.subplots_adjust(left=0.03, right=0.99)
        graphs = (hangi_setup_pc + "_" + protokol + "_" + kac_kez + '.png')
        fig.savefig(graphs)
        # plt.show()
        # plot_list_zaman = []
        plt.close(fig)

    #print(ping_sonuclar_list)
    return ping_sonuclar_list

#ping_excel()